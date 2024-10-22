import sys
import os
import logging
import time
from datetime import datetime, timedelta
from PyQt5.QtWidgets import QApplication, QMainWindow, QWidget, QVBoxLayout, QPushButton, QFileDialog, QProgressBar, QLabel, QCalendarWidget, QMessageBox
from PyQt5.QtCore import Qt, QSettings, QThread, pyqtSignal, QTimer
import openpyxl
from concurrent.futures import ThreadPoolExecutor

# Налаштування логування
logging.basicConfig(filename='app.log', level=logging.DEBUG, 
                    format='%(asctime)s - %(levelname)s - %(message)s')

class ReportGenerationThread(QThread):
    # Сигнали для оновлення прогресу, статусу та завершення
    progress = pyqtSignal(int)
    status = pyqtSignal(str)
    finished = pyqtSignal(str)

    def __init__(self, template_path, source_path, output_path, start_date):
        super().__init__()
        self.template_path = template_path
        self.source_path = source_path
        self.output_path = output_path
        self.start_date = start_date

    def run(self):
        """Основний метод, що виконується в окремому потоці для генерації звіту."""
        try:
            # Визначення кінцевої дати звіту (через 6 днів від початкової)
            end_date = self.start_date + timedelta(days=6)
            # Створення папки для збереження звіту з форматом дати "dd.mm.yyyy"
            output_folder = os.path.join(self.output_path, f"Тижневий звіт {self.start_date.strftime('%d.%m.%Y')}-{end_date.strftime('%d.%m.%Y')}")
            os.makedirs(output_folder, exist_ok=True)
            logging.info(f"Created report folder: {output_folder}")

            # Отримання списку всіх файлів шаблонів у вказаній директорії
            template_files = [f for f in os.listdir(self.template_path) if f.endswith('.xlsx')]
            total_files = len(template_files)
            self.progress.emit(0)

            # Список для зберігання дат, за які відсутні вихідні файли
            missing_files = []

            # Використання ThreadPoolExecutor для паралельної обробки шаблонів
            with ThreadPoolExecutor() as executor:
                futures = {executor.submit(self.process_template, template_file, output_folder, missing_files): template_file for template_file in template_files}
                for i, future in enumerate(futures):
                    try:
                        future.result()
                    except Exception as e:
                        logging.error(f"Error processing template {futures[future]}: {str(e)}")
                    # Оновлення прогресу
                    self.progress.emit(int((i + 1) / total_files * 100))

            # Перевірка наявності відсутніх файлів та виведення відповідного повідомлення
            if missing_files:
                missing_dates = ', '.join(sorted(set(missing_files)))
                self.finished.emit(f"Генерація звіту завершена! Відсутні вихідні файли за дати: {missing_dates}")
            else:
                self.finished.emit("Генерація звіту завершена!")
        except Exception as e:
            logging.error(f"Error during report generation: {str(e)}")
            self.finished.emit(f"Помилка: {str(e)}")

    def process_template(self, template_file, output_folder, missing_files):
        """Обробка окремого файлу шаблону."""
        self.status.emit(f"Обробка {template_file}")
        logging.debug(f"Processing template file: {template_file}")
        template_path = os.path.join(self.template_path, template_file)
        output_file = os.path.join(output_folder, template_file)
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active

        # Перебір дат тижня для пошуку та копіювання даних
        for j, date in enumerate([(self.start_date + timedelta(days=x)) for x in range(7)]):
            source_file = self.find_source_file(template_file, date)
            if source_file:
                logging.debug(f"Found source file for {template_file} on {date}: {source_file}")
                self.copy_data(source_file, ws, j)
            else:
                logging.warning(f"Source file not found for {template_file} on {date}")
                missing_files.append(date.strftime('%d.%m.%Y'))

        wb.save(output_file)
        wb.close()  # Закриття файлу після збереження
        logging.info(f"Saved report for template: {template_file}")

    def find_source_file(self, template_name, date):
        """Пошук вихідного файлу для заданого шаблону та дати."""
        try:
            for root, dirs, files in os.walk(self.source_path):
                if date.strftime('%d.%m.%Y') in root and template_name in files:
                    logging.debug(f"Source file found: {os.path.join(root, template_name)}")
                    return os.path.join(root, template_name)
        except Exception as e:
            logging.error(f"Error finding source file: {str(e)}")
        return None

    def copy_data(self, source_file, dest_worksheet, day_index):
        """Копіювання даних з вихідного файлу у шаблон."""
        try:
            logging.debug(f"Copying data from source file: {source_file}")
            source_wb = openpyxl.load_workbook(source_file, data_only=True)
            source_ws = source_wb.active

            # Копіювання даних з вихідного файлу у відповідні комірки шаблону
            for row in range(3, 28):
                dest_cell = dest_worksheet.cell(row=row+2, column=day_index+3)
                source_cell = source_ws.cell(row=row, column=3)
                dest_cell.value = source_cell.value

            source_wb.close()  # Закриття вихідного файлу після копіювання
            logging.debug(f"Data copied successfully from {source_file}")
        except Exception as e:
            logging.error(f"Error copying data from {source_file}: {str(e)}")

class SolarForecastApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Генератор прогнозу сонячної енергії")
        self.setGeometry(100, 100, 600, 400)

        self.settings = QSettings("YourCompany", "SolarForecastApp")

        self.central_widget = QWidget()
        self.setCentralWidget(self.central_widget)
        self.layout = QVBoxLayout(self.central_widget)

        self.setup_ui()

    def setup_ui(self):
        """Налаштування інтерфейсу користувача."""
        # Кнопки вибору шляхів та мітки
        self.template_path_btn = QPushButton("Вибрати шлях до шаблону")
        self.template_path_btn.clicked.connect(self.select_template_path)
        self.layout.addWidget(self.template_path_btn)

        self.template_path_label = QLabel("Шлях до шаблону: Не вибрано")
        self.layout.addWidget(self.template_path_label)

        self.source_path_btn = QPushButton("Вибрати шлях до джерела")
        self.source_path_btn.clicked.connect(self.select_source_path)
        self.layout.addWidget(self.source_path_btn)

        self.source_path_label = QLabel("Шлях до джерела: Не вибрано")
        self.layout.addWidget(self.source_path_label)

        self.output_path_btn = QPushButton("Вибрати шлях для збереження")
        self.output_path_btn.clicked.connect(self.select_output_path)
        self.layout.addWidget(self.output_path_btn)

        self.output_path_label = QLabel("Шлях для збереження: Не вибрано")
        self.layout.addWidget(self.output_path_label)

        # Календар для вибору дати
        self.calendar = QCalendarWidget()
        self.layout.addWidget(self.calendar)

        # Встановити дату за замовчуванням на завтра
        tomorrow = datetime.now() + timedelta(days=1)
        self.calendar.setSelectedDate(tomorrow)

        # Кнопка генерації звіту
        self.generate_btn = QPushButton("Згенерувати звіт")
        self.generate_btn.clicked.connect(self.generate_report)
        self.layout.addWidget(self.generate_btn)

        # Прогрес-бар
        self.progress_bar = QProgressBar()
        self.layout.addWidget(self.progress_bar)

        # Мітка статусу
        self.status_label = QLabel()
        self.layout.addWidget(self.status_label)

        # Мітка часу виконання
        self.elapsed_time_label = QLabel("Час виконання: 00:00:00")
        self.layout.addWidget(self.elapsed_time_label)

        # Таймер для оновлення часу виконання
        self.timer = QTimer()
        self.timer.timeout.connect(self.update_elapsed_time)
        self.start_time = None

        # Завантажити збережені шляхи
        self.load_saved_paths()

    def select_template_path(self):
        """Вибір директорії шаблонів."""
        path = QFileDialog.getExistingDirectory(self, "Вибрати директорію шаблону")
        if path:
            self.template_path = path
            self.settings.setValue("template_path", path)
            self.template_path_label.setText(f"Шлях до шаблону: {path}")

    def select_source_path(self):
        """Вибір директорії джерел."""
        path = QFileDialog.getExistingDirectory(self, "Вибрати директорію джерела")
        if path:
            self.source_path = path
            self.settings.setValue("source_path", path)
            self.source_path_label.setText(f"Шлях до джерела: {path}")

    def select_output_path(self):
        """Вибір директорії для збереження звітів."""
        path = QFileDialog.getExistingDirectory(self, "Вибрати директорію для збереження")
        if path:
            self.output_path = path
            self.settings.setValue("output_path", path)
            self.output_path_label.setText(f"Шлях для збереження: {path}")

    def load_saved_paths(self):
        """Завантаження збережених шляхів з налаштувань."""
        self.template_path = self.settings.value("template_path", "")
        self.source_path = self.settings.value("source_path", "")
        self.output_path = self.settings.value("output_path", "")

        # Оновити мітки збереженими шляхами
        self.template_path_label.setText(f"Шлях до шаблону: {self.template_path or 'Не вибрано'}")
        self.source_path_label.setText(f"Шлях до джерела: {self.source_path or 'Не вибрано'}")
        self.output_path_label.setText(f"Шлях для збереження: {self.output_path or 'Не вибрано'}")

    def generate_report(self):
        """Запуск процесу генерації звіту."""
        start_date = self.calendar.selectedDate().toPyDate()
        if not all([self.template_path, self.source_path, self.output_path]):
            QMessageBox.critical(self, "Помилка", "Будь ласка, вкажіть всі необхідні шляхи")
            return

        self.start_time = time.time()
        self.timer.start(1000)  # Оновлювати кожну секунду

        self.thread = ReportGenerationThread(self.template_path, self.source_path, self.output_path, start_date)
        self.thread.progress.connect(self.progress_bar.setValue)
        self.thread.status.connect(self.status_label.setText)
        self.thread.finished.connect(self.on_report_finished)
        self.thread.start()

    def update_elapsed_time(self):
        """Оновлення мітки часу виконання."""
        if self.start_time is not None:
            elapsed_time = int(time.time() - self.start_time)
            formatted_time = str(timedelta(seconds=elapsed_time))
            self.elapsed_time_label.setText(f"Час виконання: {formatted_time}")

    def on_report_finished(self, message):
        """Дії після завершення генерації звіту."""
        self.timer.stop()
        self.status_label.setText(message)
        QMessageBox.information(self, "Інформація", message)

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = SolarForecastApp()
    window.show()
    sys.exit(app.exec_())
